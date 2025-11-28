"""
Excel to Image Converter

Converts Excel spreadsheets to high-resolution images for vision model processing.
"""

import logging
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont
import io
import base64
from typing import Tuple, Optional

logger = logging.getLogger(__name__)


class ExcelToImageConverter:
    """Converts Excel sheets to images with proper formatting."""
    
    def __init__(self, cell_width: int = 150, cell_height: int = 30, 
                 font_size: int = 12, padding: int = 5):
        """
        Initialize converter with rendering parameters.
        
        Args:
            cell_width: Width of each cell in pixels
            cell_height: Height of each cell in pixels
            font_size: Font size for text
            padding: Padding inside cells
        """
        self.cell_width = cell_width
        self.cell_height = cell_height
        self.font_size = font_size
        self.padding = padding
        
        # Try to load a good font, fallback to default
        try:
            self.font = ImageFont.truetype("arial.ttf", font_size)
            self.font_bold = ImageFont.truetype("arialbd.ttf", font_size)
        except:
            try:
                # Try common Windows fonts
                self.font = ImageFont.truetype("C:/Windows/Fonts/arial.ttf", font_size)
                self.font_bold = ImageFont.truetype("C:/Windows/Fonts/arialbd.ttf", font_size)
            except:
                logger.warning("Could not load Arial font, using default")
                self.font = ImageFont.load_default()
                self.font_bold = ImageFont.load_default()
    
    def convert_excel_to_image(self, excel_path: str, sheet_name: Optional[str] = None) -> Tuple[Image.Image, str]:
        """
        Convert an Excel file to a high-resolution image with dynamic column widths.
        
        Args:
            excel_path: Path to Excel file
            sheet_name: Optional sheet name (uses first sheet if not specified)
            
        Returns:
            Tuple of (PIL Image object, base64 encoded string)
        """
        logger.info(f"Converting Excel file to image: {excel_path}")
        
        # Load workbook
        wb = load_workbook(excel_path, data_only=True)
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        # Get the used range
        min_row = ws.min_row
        max_row = ws.max_row
        min_col = ws.min_column
        max_col = ws.max_column
        
        logger.info(f"Excel range: rows {min_row}-{max_row}, cols {min_col}-{max_col}")
        
        # First pass: Calculate optimal column widths based on content
        col_widths = {}
        all_cells = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=min_row, max_row=max_row, 
                                                     min_col=min_col, max_col=max_col)):
            row_cells = []
            for col_idx, cell in enumerate(row):
                # Format the value
                value = cell.value
                if value is not None:
                    if isinstance(value, (int, float)):
                        if isinstance(value, float):
                            text = f"{value:,.2f}"
                        else:
                            text = f"{value:,}"
                    else:
                        text = str(value)
                else:
                    text = ""
                
                row_cells.append(text)
                
                # Calculate text width using font
                is_header = (row_idx == 0)
                font = self.font_bold if is_header else self.font
                
                # Get text bounding box
                try:
                    bbox = font.getbbox(text)
                    text_width = bbox[2] - bbox[0]
                except:
                    # Fallback for older Pillow versions
                    text_width = len(text) * (self.font_size // 2)
                
                # Add padding
                required_width = text_width + (2 * self.padding) + 10  # Extra 10px margin
                
                # Update column width if this cell needs more space
                if col_idx not in col_widths or required_width > col_widths[col_idx]:
                    col_widths[col_idx] = required_width
            
            all_cells.append(row_cells)
        
        # Ensure minimum width for all columns
        for col_idx in range(len(col_widths)):
            if col_widths[col_idx] < 80:  # Minimum 80px
                col_widths[col_idx] = 80
            # Cap maximum width to prevent extremely wide images
            if col_widths[col_idx] > 600:
                col_widths[col_idx] = 600
                logger.warning(f"Column {col_idx} width capped at 600px")
        
        logger.info(f"Calculated column widths: {col_widths}")
        
        # Calculate total image dimensions
        num_rows = max_row - min_row + 1
        img_width = sum(col_widths.values())
        img_height = num_rows * self.cell_height
        
        logger.info(f"Creating image: {img_width}x{img_height} pixels")
        
        # Create image with white background
        img = Image.new('RGB', (img_width, img_height), color='white')
        draw = ImageDraw.Draw(img)
        
        # Second pass: Draw cells with calculated widths
        for row_idx, row_cells in enumerate(all_cells):
            # Calculate x position for this row
            x_offset = 0
            
            for col_idx, text in enumerate(row_cells):
                col_width = col_widths[col_idx]
                y = row_idx * self.cell_height
                
                # Draw cell border
                draw.rectangle(
                    [x_offset, y, x_offset + col_width, y + self.cell_height],
                    outline='black',
                    width=1
                )
                
                # Determine if this is a header row
                is_header = (row_idx == 0)
                
                # Add light gray background for headers
                if is_header:
                    draw.rectangle(
                        [x_offset + 1, y + 1, x_offset + col_width - 1, y + self.cell_height - 1],
                        fill='#E0E0E0'
                    )
                
                # Draw text if not empty
                if text:
                    font = self.font_bold if is_header else self.font
                    text_x = x_offset + self.padding
                    text_y = y + self.padding
                    
                    # Check if text still needs truncation (for capped columns)
                    try:
                        bbox = font.getbbox(text)
                        text_width = bbox[2] - bbox[0]
                    except:
                        text_width = len(text) * (self.font_size // 2)
                    
                    available_width = col_width - (2 * self.padding)
                    
                    if text_width > available_width:
                        # Need to truncate - binary search for optimal length
                        left, right = 0, len(text)
                        truncated_text = text
                        
                        while left < right:
                            mid = (left + right + 1) // 2
                            test_text = text[:mid] + "..."
                            try:
                                bbox = font.getbbox(test_text)
                                test_width = bbox[2] - bbox[0]
                            except:
                                test_width = len(test_text) * (self.font_size // 2)
                            
                            if test_width <= available_width:
                                truncated_text = test_text
                                left = mid
                            else:
                                right = mid - 1
                        
                        text = truncated_text
                        logger.debug(f"Truncated text in col {col_idx}: {text}")
                    
                    draw.text(
                        (text_x, text_y),
                        text,
                        fill='black',
                        font=font
                    )
                
                # Move to next column
                x_offset += col_width
        
        # Convert to base64 for API transmission
        buffered = io.BytesIO()
        img.save(buffered, format="PNG", optimize=True)
        img_base64 = base64.b64encode(buffered.getvalue()).decode('utf-8')
        
        logger.info(f"Image created successfully. Size: {len(img_base64)} bytes (base64)")
        
        return img, img_base64
    
    def save_image(self, img: Image.Image, output_path: str):
        """
        Save the image to a file.
        
        Args:
            img: PIL Image object
            output_path: Path to save the image
        """
        img.save(output_path, format="PNG", optimize=True)
        logger.info(f"Image saved to: {output_path}")


def convert_excel_to_image(excel_path: str, output_path: Optional[str] = None,
                          sheet_name: Optional[str] = None) -> Tuple[Image.Image, str]:
    """
    Convenience function to convert Excel to image.
    
    Args:
        excel_path: Path to Excel file
        output_path: Optional path to save image file
        sheet_name: Optional sheet name
        
    Returns:
        Tuple of (PIL Image object, base64 encoded string)
    """
    converter = ExcelToImageConverter()
    img, img_base64 = converter.convert_excel_to_image(excel_path, sheet_name)
    
    if output_path:
        converter.save_image(img, output_path)
    
    return img, img_base64

